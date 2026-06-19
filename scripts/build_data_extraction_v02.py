import pathlib

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTDIR = pathlib.Path("04_data_extraction")
DEFAULT_OUTPUT_DIR = pathlib.Path(
    r"D:\001综合文件夹海康备份\44 自己的系列\52自己的写作文章\02肺栓塞和DVT\02肺栓塞网状meta-cdex版"
)


def row(
    study_id,
    short_name,
    arm_label,
    node,
    n,
    risk,
    death_7d=None,
    death_30d=None,
    clinical_deterioration=None,
    primary_composite=None,
    major_bleeding=None,
    ich=None,
    recurrent_pe=None,
    rv_lv=None,
    follow_up=None,
    locked="partial",
    ready_death="no",
    ready_clinical="no",
    ready_bleeding="no",
    ready_ich="no",
    source_note="",
):
    return {
        "study_id": study_id,
        "short_name": short_name,
        "arm_label": arm_label,
        "node": node,
        "n_randomized_or_analyzed": n,
        "risk_mapping_v0.2": risk,
        "death_7d_n": death_7d,
        "death_followup_n": death_30d,
        "clinical_deterioration_n": clinical_deterioration,
        "primary_clinical_composite_n": primary_composite,
        "major_bleeding_n": major_bleeding,
        "major_bleeding_definition": "",
        "intracranial_hemorrhage_n": ich,
        "recurrent_pe_n": recurrent_pe,
        "rv_lv_or_rv_recovery_locked": rv_lv,
        "follow_up_locked": follow_up,
        "locked_status_v0.2": locked,
        "ready_death_nma": ready_death,
        "ready_clinical_deterioration_nma": ready_clinical,
        "ready_major_bleeding_nma": ready_bleeding,
        "ready_ich_nma": ready_ich,
        "source_note_v0.2": source_note,
    }


rows = [
    row("TOPCOAT_2014", "TOPCOAT", "Tenecteplase", "ST", 40, "intermediate-risk; mixed low/intermediate-high features",
        death_30d=1, clinical_deterioration=1, primary_composite=6, major_bleeding=1, ich=1,
        follow_up="5 days and 90 days", locked="core locked; outcome harmonization needed",
        ready_death="yes", ready_clinical="sensitivity", ready_bleeding="yes", ready_ich="yes",
        source_note="Results: 1 tenecteplase patient died from ICH 5h after drug; 6/40 had 90-day adverse outcome; only major bleed during 5-day surveillance."),
    row("TOPCOAT_2014", "TOPCOAT", "Placebo + anticoagulation", "AC", 43, "intermediate-risk; mixed low/intermediate-high features",
        death_30d=1, clinical_deterioration=3, primary_composite=16, major_bleeding=0, ich=0,
        follow_up="5 days and 90 days", locked="core locked; outcome harmonization needed",
        ready_death="yes", ready_clinical="sensitivity", ready_bleeding="yes", ready_ich="yes",
        source_note="Results: 3 early adverse outcomes in placebo, including 1 PE-attributed death and 2 shock/intubation/thrombectomy; total 16/43 adverse outcome by 90 days."),

    row("ULTIMA_2014", "ULTIMA", "USAT/USCDT + UFH", "USCDT", 30, "intermediate-risk by RV/LV >=1; biomarker not required",
        death_30d=0, clinical_deterioration=0, major_bleeding=0, ich=0, recurrent_pe=0,
        rv_lv="1.28+/-0.19 to 0.99+/-0.17 at 24h; change 0.30+/-0.20",
        follow_up="90 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Abstract/results: safety outcomes at 90 days: 0 deaths in USAT, no major bleeding, no recurrent VTE."),
    row("ULTIMA_2014", "ULTIMA", "UFH alone", "AC", 29, "intermediate-risk by RV/LV >=1; biomarker not required",
        death_30d=1, clinical_deterioration=0, major_bleeding=0, ich=0, recurrent_pe=0,
        rv_lv="1.20+/-0.14 to 1.17+/-0.20 at 24h; change 0.03+/-0.16",
        follow_up="90 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Abstract/results: 1 death in heparin group, no major bleeding, no recurrent VTE."),

    row("PEITHO_2014", "PEITHO", "Tenecteplase + heparin", "ST", 506, "intermediate-high PE: RV dysfunction + positive troponin",
        death_7d=6, death_30d=12, clinical_deterioration=8, primary_composite=13, major_bleeding=58, ich=10,
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: primary 13, death 6, hemodynamic decompensation 8; Table 4/text: ISTH major bleeding 58, major extracranial 32, hemorrhagic stroke/ICH 10."),
    row("PEITHO_2014", "PEITHO", "Placebo + heparin", "AC", 499, "intermediate-high PE: RV dysfunction + positive troponin",
        death_7d=9, death_30d=16, clinical_deterioration=25, primary_composite=28, major_bleeding=12, ich=1,
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: primary 28, death 9, hemodynamic decompensation 25; Table 4/text: ISTH major bleeding 12, major extracranial 6, hemorrhagic stroke/ICH 1."),

    row("SUNSET_sPE_2021", "SUNSET sPE", "USAT/EKOS", "USCDT", 40, "submassive/intermediate PE planned for catheter lysis",
        death_30d=1, major_bleeding=2, ich=1, rv_lv="1.5+/-0.3 to 1.2+/-0.2; delta 0.35+/-0.34",
        follow_up="48h and 90 days", locked="partial locked; table assigns adverse events by group as extracted",
        ready_death="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Safety table: major bleeding events 2, minor bleeding 3, in-hospital death 1; text reports one stroke and one vaginal bleed requiring transfusion."),
    row("SUNSET_sPE_2021", "SUNSET sPE", "Standard CDT", "CDT", 41, "submassive/intermediate PE planned for catheter lysis",
        death_30d=0, major_bleeding=0, ich=0, rv_lv="1.7+/-0.4 to 1.1+/-0.2; delta 0.59+/-0.42",
        follow_up="48h and 90 days", locked="partial locked; table assigns adverse events by group as extracted",
        ready_death="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Safety table: comparator adverse events 0 in visible extraction; primary endpoint imaging only."),

    row("CANARY_2022", "CANARY", "cCDT + anticoagulation", "CDT", 48, "intermediate-high PE",
        death_30d=0, primary_composite=2, major_bleeding=1, ich=None,
        rv_lv="72h RV/LV>0.9: 13/48; 3mo RV/LV>0.9: 2/46; unrecovered RV 3/46",
        follow_up="72h and 3 months", locked="core locked for death/bleeding/RV; ICH not separately reported",
        ready_death="yes", ready_clinical="sensitivity", ready_bleeding="yes", ready_ich="no",
        source_note="Table 2: 3mo all-cause mortality 0/48; composite mortality or primary outcome 2/48; BARC 3/5 1/48."),
    row("CANARY_2022", "CANARY", "Anticoagulation", "AC", 46, "intermediate-high PE",
        death_30d=3, primary_composite=8, major_bleeding=0, ich=None,
        rv_lv="72h RV/LV>0.9: 24/46; 3mo RV/LV>0.9: 5/39; unrecovered RV 11/39",
        follow_up="72h and 3 months", locked="core locked for death/bleeding/RV; ICH not separately reported",
        ready_death="yes", ready_clinical="sensitivity", ready_bleeding="yes", ready_ich="no",
        source_note="Table 2: 3mo all-cause mortality 3/46; composite mortality or primary outcome 8/46; BARC 3/5 0."),

    row("PEERLESS_2025", "PEERLESS", "FlowTriever LBMT", "LBAT", 274, "intermediate-risk, mostly intermediate-high",
        death_7d=0, death_30d=1, clinical_deterioration=5, major_bleeding=19, ich=2,
        rv_lv="Mean RV/LV reduction 0.32+/-0.24 at 24h",
        follow_up="discharge/7 days and 30 days", locked="core locked; CDT comparator is mixed",
        ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Text: mortality 0.0% vs 0.4%, ICH 0.7% vs 0.4%, major bleeding 6.9% both; 5 vs 15 clinical deterioration/bailout; 30d mortality 0.4%."),
    row("PEERLESS_2025", "PEERLESS", "CDT mixed", "CDT", 276, "intermediate-risk, mostly intermediate-high",
        death_7d=1, death_30d=2, clinical_deterioration=15, major_bleeding=19, ich=1,
        rv_lv="Mean RV/LV reduction 0.30+/-0.26 at 24h",
        follow_up="discharge/7 days and 30 days", locked="core locked; CDT arm mixed: 59.8% USCDT, 23.2% side-hole CDT, 8.7% side-slit CDT",
        ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Text: CDT arm composition; clinical deterioration/bailout 15; major bleeding 19; 30d mortality 0.8%."),

    row("STORM_PE_2025", "STORM-PE", "CAVT/Penumbra + AC", "CAT", 47, "intermediate-high PE",
        death_7d=2, clinical_deterioration=1, primary_composite=2, major_bleeding=1, ich=None, recurrent_pe=0,
        rv_lv="Baseline 1.63+/-0.36; 48h RV/LV reduction 0.52+/-0.37",
        follow_up="48h, 7 days, 90 days planned", locked="core locked; ICH not separately stated",
        ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="no",
        source_note="Text: MAE 2/47; clinical deterioration 1; two PE-related deaths; major bleeding 1; recurrent PE 0."),
    row("STORM_PE_2025", "STORM-PE", "AC alone", "AC", 53, "intermediate-high PE",
        death_7d=0, clinical_deterioration=3, primary_composite=4, major_bleeding=1, ich=None, recurrent_pe=0,
        rv_lv="Baseline 1.56+/-0.35; 48h RV/LV reduction 0.24+/-0.40",
        follow_up="48h, 7 days, 90 days planned", locked="core locked; ICH not separately stated",
        ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="no",
        source_note="Text: MAE 4/53; clinical deterioration requiring rescue 3; major bleeding 1; recurrent PE 0."),

    row("STRATIFY_2026", "STRATIFY", "USAT low-dose alteplase", "USCDT", 70, "intermediate-risk; exact low/intermediate-high mapping pending",
        follow_up="48-96h and 3 months", locked="pending full result table extraction",
        source_note="Text confirms three arms n=70 each; clinical/safety event counts require table extraction/OCR or manual table review."),
    row("STRATIFY_2026", "STRATIFY", "IV low-dose alteplase", "ST", 70, "intermediate-risk; low-dose ST sensitivity",
        follow_up="48-96h and 3 months", locked="pending full result table extraction",
        source_note="Low-dose IV alteplase arm retained as ST-low-dose/sensitivity field."),
    row("STRATIFY_2026", "STRATIFY", "Heparin alone", "AC", 70, "intermediate-risk; exact low/intermediate-high mapping pending",
        follow_up="48-96h and 3 months", locked="pending full result table extraction",
        source_note="Pending exact death/clinical deterioration/bleeding table extraction."),

    row("PRETHA_2026", "PRETHA", "Penumbra Indigo CAT8 thrombectomy", "CAT", 13, "intermediate-high PE",
        death_30d=0, major_bleeding=0, ich=0, recurrent_pe=3,
        rv_lv="RV/LV 1.2 to 0.89 at 48h; mean decrease 0.3",
        follow_up="48h, 1/6/12 months", locked="partial locked; clinical event definition differs",
        ready_death="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Results: 13 per arm; thrombectomy had 1 non-significant hemoptysis WHO grade 2, no major/ICH; baseline table lists recurrent PE after current episode=3."),
    row("PRETHA_2026", "PRETHA", "Trans-catheter thrombolysis", "CDT", 13, "intermediate-high PE",
        death_30d=1, major_bleeding=1, ich=1, recurrent_pe=0,
        rv_lv="RV/LV 1.27 to 0.88 at 48h; mean decrease 0.4",
        follow_up="48h, 1/6/12 months", locked="partial locked; bleeding definition WHO grade",
        ready_death="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Results: 3/13 groin hematomas; 1 WHO grade 3 transfusion; 1 patient died from brain hemorrhage WHO grade 4."),
    row("PRETHA_2026", "PRETHA", "Conservative treatment", "AC", 13, "intermediate-high PE",
        death_30d=0, major_bleeding=0, ich=0, recurrent_pe=2,
        rv_lv="RV/LV 1.23 to 1.12 at 48h",
        follow_up="48h, 1/6/12 months", locked="partial locked; clinical event definition differs",
        ready_death="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Results: 1 pneumonia; no bleeding signal reported; baseline table lists recurrent PE after current episode=2."),

    row("HI_PEITHO_2026", "HI-PEITHO", "USCDT + AC", "USCDT", 273, "acute intermediate-risk PE with cardiorespiratory distress",
        death_7d=3, death_30d=5, clinical_deterioration=10, primary_composite=11, major_bleeding=11, ich=0, recurrent_pe=1,
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 2: primary 11, PE death 3, decompensation 10, recurrent PE 1; Table 3: ISTH major bleeding 11 by 30d, ICH 0; Table 4: all-cause death 5."),
    row("HI_PEITHO_2026", "HI-PEITHO", "AC alone", "AC", 271, "acute intermediate-risk PE with cardiorespiratory distress",
        death_7d=2, death_30d=3, clinical_deterioration=28, primary_composite=28, major_bleeding=8, ich=0, recurrent_pe=2,
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 2: primary 28, PE death 1, decompensation 28, recurrent PE 1; Table 3: ISTH major bleeding 8 by 30d, ICH 0; Table 4: all-cause death 3."),

    row("MAPPET3_2002", "MAPPET-3", "Heparin + alteplase", "ST", 118, "submassive PE; intermediate-risk by RV dysfunction/PH/ECG strain, biomarker not required",
        death_30d=4, clinical_deterioration=12, primary_composite=13, major_bleeding=1, ich=0, recurrent_pe=4,
        follow_up="in-hospital or day 30", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 2: primary 13, death 4, escalation 12, recurrent PE 4, major bleeding 1, hemorrhagic stroke 0."),
    row("MAPPET3_2002", "MAPPET-3", "Heparin + placebo", "AC", 138, "submassive PE; intermediate-risk by RV dysfunction/PH/ECG strain, biomarker not required",
        death_30d=3, clinical_deterioration=34, primary_composite=34, major_bleeding=5, ich=0, recurrent_pe=4,
        follow_up="in-hospital or day 30", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 2: primary 34, death 3, escalation 34, recurrent PE 4, major bleeding 5, hemorrhagic stroke 0."),

    row("HAIRE_1993", "Haire 1993", "Alteplase", "ST", None, "pending OCR/manual verification", locked="OCR_needed",
        source_note="Extracted text only shows ProQuest copyright/title page; outcome data not machine-readable."),
    row("HAIRE_1993", "Haire 1993", "Heparin", "AC", None, "pending OCR/manual verification", locked="OCR_needed",
        source_note="Needs rendered page review or OCR before inclusion."),

    row("CDT_PILOT_2022", "EuroIntervention pilot CDT", "CDT alteplase 20 mg", "CDT", 12, "intermediate-high PE",
        major_bleeding=0, ich=0, follow_up="24/48/72h",
        locked="partial locked; primary composite exact count pending", ready_bleeding="yes", ready_ich="yes",
        source_note="Abstract/results: no intracranial or life-threatening bleeding; one local BARC 2 access hematoma in CDT arm; exact clinical efficacy counts pending."),
    row("CDT_PILOT_2022", "EuroIntervention pilot CDT", "Standard anticoagulation", "AC", 11, "intermediate-high PE",
        major_bleeding=0, ich=0, follow_up="24/48/72h",
        locked="partial locked; primary composite exact count pending", ready_bleeding="yes", ready_ich="yes",
        source_note="Abstract/results: no intracranial or life-threatening bleeding; bleeding count not major/life-threatening."),

    row("FASULLO_2011", "Fasullo 2011", "Thrombolysis + heparin", "ST", 37, "submassive PE with RVD; biomarker mixed/unclear",
        death_30d=0, clinical_deterioration=0, major_bleeding=2, ich=0, recurrent_pe=0,
        rv_lv="RV function improved earlier; detailed echo table locked in source",
        follow_up="10 days and 180 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Tables 4/6: in-hospital death 0, recurrent PE 0, major bleeding 2; total death 0; text: no intracranial bleeding in thrombolysis."),
    row("FASULLO_2011", "Fasullo 2011", "Heparin/placebo", "AC", 35, "submassive PE with RVD; biomarker mixed/unclear",
        death_30d=5, clinical_deterioration=2, major_bleeding=1, ich=0, recurrent_pe=4,
        follow_up="10 days and 180 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 4: death 5, recurrent fatal PE 3, irreversible RVD 2, major bleeding 1; Table 6 total death 6 and recurrent fatal PE 4 through 180d."),

    row("SINHA_2017", "Sinha 2017", "Tenecteplase + UFH", "ST", 45, "submassive/intermediate-risk; RV dysfunction and/or biomarkers; all had myocardial injury",
        death_7d=2, death_30d=2, clinical_deterioration=2, primary_composite=2, major_bleeding=1, ich=1, recurrent_pe=2,
        rv_lv="Baseline RV/LV 1.14+/-0.11; RV function improvement 31/45",
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: primary 2, death 2, hemodynamic decompensation 2, recurrent PE 2, major bleeding 1, hemorrhagic stroke 1."),
    row("SINHA_2017", "Sinha 2017", "Placebo + UFH", "AC", 41, "submassive/intermediate-risk; RV dysfunction and/or biomarkers; all had myocardial injury",
        death_7d=2, death_30d=2, clinical_deterioration=8, primary_composite=8, major_bleeding=1, ich=0, recurrent_pe=1,
        rv_lv="Baseline RV/LV 1.16+/-0.14; RV function improvement 16/41",
        follow_up="7 and 30 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: primary 8, death 2, hemodynamic decompensation 8, recurrent PE 1, major bleeding 1, hemorrhagic stroke 0."),

    row("ZHANG_2018", "Zhang 2018", "Low-dose rt-PA + LMWH", "ST", 33, "acute intermediate-risk PE; low-dose ST sensitivity",
        death_30d=0, clinical_deterioration=0, major_bleeding=0, ich=0, recurrent_pe=1,
        rv_lv="1.26+/-0.22 to 0.96+/-0.18 at 24h",
        follow_up="90 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Safety outcomes: at 90 days no deaths and no major bleeding among 66; no hemodynamic decompensation and 1 recurrent VTE in rt-PA group."),
    row("ZHANG_2018", "Zhang 2018", "LMWH", "AC", 33, "acute intermediate-risk PE",
        death_30d=0, clinical_deterioration=3, major_bleeding=0, ich=0, recurrent_pe=2,
        rv_lv="1.22+/-0.19 to 1.17+/-0.21 at 24h",
        follow_up="90 days", locked="core locked", ready_death="yes", ready_clinical="yes", ready_bleeding="yes", ready_ich="yes",
        source_note="Safety outcomes: 3 hemodynamic decompensation and 2 recurrent VTE in LMWH group; no deaths and no major bleeding among 66."),

    row("MOPETT_2013", "MOPETT", "Safe-dose tPA + AC", "ST", 61, "moderate PE; risk mapping pending, sensitivity-only unless maps to intermediate-risk",
        death_30d=1, primary_composite=1, major_bleeding=0, ich=0, recurrent_pe=0,
        rv_lv="PASP endpoint; PH 9/58 at 28 months",
        follow_up="28+/-5 months", locked="core locked; sensitivity-only risk stratum",
        ready_death="sensitivity", ready_clinical="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Table 3: recurrent PE 0, total mortality 1, mortality+recurrent PE 1, bleeding 0."),
    row("MOPETT_2013", "MOPETT", "AC alone", "AC", 60, "moderate PE; risk mapping pending, sensitivity-only unless maps to intermediate-risk",
        death_30d=3, primary_composite=6, major_bleeding=0, ich=0, recurrent_pe=3,
        rv_lv="PASP endpoint; PH 32/56 at 28 months",
        follow_up="28+/-5 months", locked="core locked; sensitivity-only risk stratum",
        ready_death="sensitivity", ready_clinical="sensitivity", ready_bleeding="sensitivity", ready_ich="sensitivity",
        source_note="Table 3: recurrent PE 3, total mortality 3, mortality+recurrent PE 6, bleeding 0."),

    row("AHMED_2018", "Ahmed 2018", "Streptokinase + anticoagulation", "ST", 24, "submassive PE causing RV dysfunction; biomarker criteria stated in discussion",
        death_30d=None, major_bleeding=0, ich=0,
        rv_lv="PASP 51.54+/-6.95 to 39.08+/-3.83 at 72h; PH endpoint 12/24",
        follow_up="72h; hospital stay", locked="partial locked; mortality not reported as event count",
        ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: endpoint pulmonary hypertension 12/24; no bleeding detected in both groups; death count not explicitly reported."),
    row("AHMED_2018", "Ahmed 2018", "Anticoagulation", "AC", 28, "submassive PE causing RV dysfunction; biomarker criteria stated in discussion",
        death_30d=None, major_bleeding=0, ich=0,
        rv_lv="PASP 47.92+/-8.02 to 46.5+/-7.22 at 72h; PH endpoint 24/28",
        follow_up="72h; hospital stay", locked="partial locked; mortality not reported as event count",
        ready_bleeding="yes", ready_ich="yes",
        source_note="Table 3: endpoint pulmonary hypertension 24/28; no bleeding detected in both groups; death count not explicitly reported."),
]


for r in rows:
    if r["study_id"] in {"PEITHO_2014", "HI_PEITHO_2026"}:
        r["major_bleeding_definition"] = "ISTH major bleeding"
    elif r["study_id"] == "CANARY_2022":
        r["major_bleeding_definition"] = "BARC type 3 or 5"
    elif r["study_id"] == "STORM_PE_2025":
        r["major_bleeding_definition"] = "BARC major bleeding"
    elif r["study_id"] == "CDT_PILOT_2022":
        r["major_bleeding_definition"] = "No intracranial/life-threatening BARC type 5 or 3c bleeding"
    elif r["study_id"] == "PRETHA_2026":
        r["major_bleeding_definition"] = "WHO bleeding grade; sensitivity only"
    else:
        r["major_bleeding_definition"] = "Author-defined major bleeding"


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    columns = [str(c) for c in df.columns]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for values in df.itertuples(index=False, name=None):
        lines.append("| " + " | ".join("" if pd.isna(v) else str(v).replace("\n", " ") for v in values) + " |")
    return "\n".join(lines)


def style_excel(path: pathlib.Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20 if idx < 7 else 28
        ws.column_dimensions["U"].width = 80
    wb.save(path)


def main() -> None:
    OUTDIR.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(rows)
    summary = (
        df.groupby("study_id", dropna=False)
        .agg(
            short_name=("short_name", "first"),
            arms=("arm_label", "count"),
            nodes=("node", lambda x: " vs ".join(x.astype(str))),
            locked_status=("locked_status_v0.2", lambda x: "; ".join(sorted(set(x.astype(str))))),
            ready_death=("ready_death_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_clinical=("ready_clinical_deterioration_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_bleeding=("ready_major_bleeding_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_ich=("ready_ich_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
        )
        .reset_index()
    )
    audit = pd.DataFrame(
        [
            {"item": "studies_with_at_least_partial_lock", "value": df.loc[df["locked_status_v0.2"] != "OCR_needed", "study_id"].nunique()},
            {"item": "arm_rows", "value": len(df)},
            {"item": "death_nma_ready_or_sensitivity_arms", "value": df["ready_death_nma"].isin(["yes", "sensitivity"]).sum()},
            {"item": "clinical_nma_ready_or_sensitivity_arms", "value": df["ready_clinical_deterioration_nma"].isin(["yes", "sensitivity"]).sum()},
            {"item": "major_bleeding_ready_or_sensitivity_arms", "value": df["ready_major_bleeding_nma"].isin(["yes", "sensitivity"]).sum()},
            {"item": "ich_ready_or_sensitivity_arms", "value": df["ready_ich_nma"].isin(["yes", "sensitivity"]).sum()},
            {"item": "do_not_analyze_without_manual_step", "value": "HAIRE_1993 requires OCR/manual review; STRATIFY_2026 requires full result-table extraction."},
        ]
    )

    xlsx_path = OUTDIR / "data_extraction_v0.2_locked_core_outcomes.xlsx"
    arm_csv = OUTDIR / "data_extraction_arms_v0.2_locked_core_outcomes.csv"
    summary_csv = OUTDIR / "data_extraction_study_summary_v0.2_locked_core_outcomes.csv"
    md_path = OUTDIR / "data_extraction_v0.2_locked_core_outcomes.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Arm_Level_v0.2", index=False)
        summary.to_excel(writer, sheet_name="Study_Summary_v0.2", index=False)
        audit.to_excel(writer, sheet_name="Audit", index=False)
    style_excel(xlsx_path)
    df.to_csv(arm_csv, index=False, encoding="utf-8-sig")
    summary.to_csv(summary_csv, index=False, encoding="utf-8-sig")
    md_path.write_text(
        "\n".join(
            [
                "# Data Extraction v0.2 Locked Core Outcomes",
                "",
                "This v0.2 file converts v0.1 into an analysis-oriented dataset for core outcomes. Rows marked `pending full result table extraction`, `OCR_needed`, or `sensitivity` must not be used in the primary NMA without the specified caveat.",
                "",
                "## Audit",
                "",
                dataframe_to_markdown(audit),
                "",
                "## Study Summary",
                "",
                dataframe_to_markdown(summary),
                "",
                "## Arm-Level Dataset",
                "",
                dataframe_to_markdown(df),
                "",
            ]
        ),
        encoding="utf-8",
    )

    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    copies = {
        xlsx_path: "正式数据提取表_v0.2_核心结局锁定.xlsx",
        arm_csv: "正式数据提取表_治疗臂层_v0.2_核心结局锁定.csv",
        summary_csv: "正式数据提取表_研究汇总_v0.2_核心结局锁定.csv",
        md_path: "正式数据提取表_v0.2_核心结局锁定.md",
        OUTDIR / "fulltext_verification_snippets_v0.2.txt": "全文核验关键片段_v0.2.txt",
    }
    for src, dest_name in copies.items():
        if src.exists():
            (DEFAULT_OUTPUT_DIR / dest_name).write_bytes(src.read_bytes())

    print(f"WROTE\t{xlsx_path}")
    print(f"WROTE\t{arm_csv}")
    print(f"WROTE\t{summary_csv}")
    print(f"WROTE\t{md_path}")
    print(f"STUDIES\t{df['study_id'].nunique()}")
    print(f"ARMS\t{len(df)}")
    print(f"LOCKED_OR_PARTIAL_STUDIES\t{df.loc[df['locked_status_v0.2'] != 'OCR_needed', 'study_id'].nunique()}")
    print(f"COPIED_TO\t{DEFAULT_OUTPUT_DIR}")


if __name__ == "__main__":
    main()
