from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "04_data_extraction"


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    columns = [str(c) for c in df.columns]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for values in df.itertuples(index=False, name=None):
        lines.append("| " + " | ".join("" if pd.isna(v) else str(v).replace("\n", " ") for v in values) + " |")
    return "\n".join(lines)


def style_excel(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22 if idx < 8 else 30
        if ws.max_column >= 22:
            ws.column_dimensions["V"].width = 88
    wb.save(path)


def update_core_outcomes() -> None:
    input_csv = DATA_DIR / "data_extraction_arms_v0.2_locked_core_outcomes.csv"
    df = pd.read_csv(input_csv)

    updates = {
        "USAT low-dose alteplase": {
            "node": "USCDT",
            "n_randomized_or_analyzed": 71,
            "risk_mapping_v0.2": "intermediate-high PE by ESC criteria",
            "death_7d_n": 2,
            "death_followup_n": 2,
            "clinical_deterioration_n": 1,
            "primary_clinical_composite_n": None,
            "major_bleeding_n": 6,
            "major_bleeding_definition": "ISTH major bleeding; TIMI major also reported",
            "intracranial_hemorrhage_n": None,
            "recurrent_pe_n": None,
            "rv_lv_or_rv_recovery_locked": "RV/LV diameter ratio reduction 48-96h: 0.3 +/- 0.3",
            "follow_up_locked": "48-96h imaging and 3 months",
            "locked_status_v0.2": "core locked for STRATIFY v0.3; ICH arm-specific not reported",
            "ready_death_nma": "yes",
            "ready_clinical_deterioration_nma": "yes",
            "ready_major_bleeding_nma": "yes",
            "ready_ich_nma": "no",
            "source_note_v0.2": (
                "STRATIFY Table 2/Figure 1: randomized n=71; deaths 2 by 3 months; "
                "full-dose thrombolysis for clinical deterioration 1; ISTH major bleeding 6; "
                "RV/LV reduction 0.3 SD 0.3 at 48-96h. Text confirms two fatal intracranial "
                "haemorrhages overall but does not provide complete arm-specific ICH counts."
            ),
        },
        "IV low-dose alteplase": {
            "node": "ST",
            "n_randomized_or_analyzed": 70,
            "risk_mapping_v0.2": "intermediate-high PE by ESC criteria; low-dose ST sensitivity",
            "death_7d_n": 0,
            "death_followup_n": 4,
            "clinical_deterioration_n": 2,
            "primary_clinical_composite_n": None,
            "major_bleeding_n": 6,
            "major_bleeding_definition": "ISTH major bleeding; TIMI major also reported",
            "intracranial_hemorrhage_n": None,
            "recurrent_pe_n": None,
            "rv_lv_or_rv_recovery_locked": "RV/LV diameter ratio reduction 48-96h: 0.4 +/- 0.4",
            "follow_up_locked": "48-96h imaging and 3 months",
            "locked_status_v0.2": "core locked for STRATIFY v0.3; ICH arm-specific not reported",
            "ready_death_nma": "yes",
            "ready_clinical_deterioration_nma": "yes",
            "ready_major_bleeding_nma": "yes",
            "ready_ich_nma": "no",
            "source_note_v0.2": (
                "STRATIFY Table 2/Figure 1: randomized n=70; deaths 4 by 3 months; "
                "full-dose thrombolysis for clinical deterioration 2; ISTH major bleeding 6; "
                "RV/LV reduction 0.4 SD 0.4 at 48-96h. Low-dose IV alteplase should be flagged "
                "as ST low-dose/sensitivity."
            ),
        },
        "Heparin alone": {
            "node": "AC",
            "n_randomized_or_analyzed": 69,
            "risk_mapping_v0.2": "intermediate-high PE by ESC criteria",
            "death_7d_n": 0,
            "death_followup_n": 0,
            "clinical_deterioration_n": 2,
            "primary_clinical_composite_n": None,
            "major_bleeding_n": 1,
            "major_bleeding_definition": "ISTH major bleeding; TIMI major also reported",
            "intracranial_hemorrhage_n": None,
            "recurrent_pe_n": None,
            "rv_lv_or_rv_recovery_locked": "RV/LV diameter ratio reduction 48-96h: 0.3 +/- 0.3",
            "follow_up_locked": "48-96h imaging and 3 months",
            "locked_status_v0.2": "core locked for STRATIFY v0.3; ICH arm-specific not reported",
            "ready_death_nma": "yes",
            "ready_clinical_deterioration_nma": "yes",
            "ready_major_bleeding_nma": "yes",
            "ready_ich_nma": "no",
            "source_note_v0.2": (
                "STRATIFY Table 2/Figure 1: randomized n=69; deaths 0 by 3 months; "
                "full-dose thrombolysis for clinical deterioration 2; ISTH major bleeding 1; "
                "RV/LV reduction 0.3 SD 0.3 at 48-96h."
            ),
        },
    }

    for arm_label, values in updates.items():
        mask = (df["study_id"] == "STRATIFY_2026") & (df["arm_label"] == arm_label)
        if mask.sum() != 1:
            raise RuntimeError(f"Expected one STRATIFY row for {arm_label}, found {mask.sum()}")
        for key, value in values.items():
            df.loc[mask, key] = value

    summary = (
        df.groupby("study_id", dropna=False)
        .agg(
            short_name=("short_name", "first"),
            arms=("arm_label", "count"),
            nodes=("node", lambda x: " vs ".join(x.astype(str))),
            locked_status=("locked_status_v0.2", lambda x: "; ".join(sorted(set(x.astype(str))))),
            ready_death=("ready_death_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_clinical=("ready_clinical_deterioration_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_bleeding=("ready_major_bleeding_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
            ready_ich=("ready_ich_nma", lambda x: ", ".join(sorted(set(x.astype(str))))),
        )
        .reset_index()
    )
    audit = pd.DataFrame(
        [
            {"item": "version", "value": "v0.3_STRATIFY_locked"},
            {"item": "arm_rows", "value": len(df)},
            {"item": "studies", "value": df["study_id"].nunique()},
            {"item": "stratify_change", "value": "STRATIFY event counts and RV/LV values locked from Table 2/Figure 1"},
            {"item": "remaining_manual_step", "value": "HAIRE_1993 OCR/manual review; STRATIFY arm-specific ICH unavailable in article tables"},
        ]
    )

    xlsx_path = DATA_DIR / "data_extraction_v0.3_STRATIFY_locked.xlsx"
    arm_csv = DATA_DIR / "data_extraction_arms_v0.3_STRATIFY_locked.csv"
    summary_csv = DATA_DIR / "data_extraction_study_summary_v0.3_STRATIFY_locked.csv"
    md_path = DATA_DIR / "data_extraction_v0.3_STRATIFY_locked.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Arm_Level_v0.3", index=False)
        summary.to_excel(writer, sheet_name="Study_Summary_v0.3", index=False)
        audit.to_excel(writer, sheet_name="Audit", index=False)
    style_excel(xlsx_path)
    df.to_csv(arm_csv, index=False, encoding="utf-8-sig")
    summary.to_csv(summary_csv, index=False, encoding="utf-8-sig")
    md_path.write_text(
        "\n".join(
            [
                "# Data Extraction v0.3 STRATIFY Locked",
                "",
                "Incremental update from v0.2. STRATIFY 2026 was checked against full-text Table 2 and Figure 1.",
                "",
                "## Audit",
                "",
                dataframe_to_markdown(audit),
                "",
                "## STRATIFY Rows",
                "",
                dataframe_to_markdown(df[df["study_id"] == "STRATIFY_2026"]),
                "",
                "## Study Summary",
                "",
                dataframe_to_markdown(summary),
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(f"WROTE {xlsx_path}")
    print(f"WROTE {arm_csv}")
    print(f"WROTE {summary_csv}")
    print(f"WROTE {md_path}")


def update_rv_lv() -> None:
    input_csv = DATA_DIR / "RV_LV_data_v0.1.csv"
    df = pd.read_csv(input_csv)
    df = df[df["study_id"] != "STRATIFY_2026"].copy()
    additions = pd.DataFrame(
        [
            {
                "study_id": "STRATIFY_2026",
                "short_name": "STRATIFY",
                "node": "USCDT",
                "arm_label": "USAT low-dose alteplase",
                "n": 71,
                "baseline_mean": "",
                "baseline_sd": "",
                "early_timepoint": "48-96h",
                "early_followup_mean": "",
                "early_followup_sd": "",
                "early_change_reduction_mean": 0.3,
                "early_change_reduction_sd": 0.3,
                "long_term_timepoint": "3mo",
                "long_term_value": "",
                "rv_recovery_event_n": "",
                "rv_recovery_total_n": "",
                "measure": "CT RV/LV diameter ratio reduction",
                "analysis_status": "main_continuous_nma",
                "conversion_needed": "no",
                "source_note": "STRATIFY Table 2: RV/LV diameter ratio reduction 0.3 +/- 0.3 at 48-96h.",
            },
            {
                "study_id": "STRATIFY_2026",
                "short_name": "STRATIFY",
                "node": "ST",
                "arm_label": "IV low-dose alteplase",
                "n": 70,
                "baseline_mean": "",
                "baseline_sd": "",
                "early_timepoint": "48-96h",
                "early_followup_mean": "",
                "early_followup_sd": "",
                "early_change_reduction_mean": 0.4,
                "early_change_reduction_sd": 0.4,
                "long_term_timepoint": "3mo",
                "long_term_value": "",
                "rv_recovery_event_n": "",
                "rv_recovery_total_n": "",
                "measure": "CT RV/LV diameter ratio reduction",
                "analysis_status": "main_continuous_nma",
                "conversion_needed": "no",
                "source_note": "STRATIFY Table 2: RV/LV diameter ratio reduction 0.4 +/- 0.4 at 48-96h; low-dose ST sensitivity flag.",
            },
            {
                "study_id": "STRATIFY_2026",
                "short_name": "STRATIFY",
                "node": "AC",
                "arm_label": "Heparin alone",
                "n": 69,
                "baseline_mean": "",
                "baseline_sd": "",
                "early_timepoint": "48-96h",
                "early_followup_mean": "",
                "early_followup_sd": "",
                "early_change_reduction_mean": 0.3,
                "early_change_reduction_sd": 0.3,
                "long_term_timepoint": "3mo",
                "long_term_value": "",
                "rv_recovery_event_n": "",
                "rv_recovery_total_n": "",
                "measure": "CT RV/LV diameter ratio reduction",
                "analysis_status": "main_continuous_nma",
                "conversion_needed": "no",
                "source_note": "STRATIFY Table 2: RV/LV diameter ratio reduction 0.3 +/- 0.3 at 48-96h.",
            },
        ]
    )
    df = pd.concat([df, additions], ignore_index=True)
    summary = (
        df.groupby(["analysis_status", "conversion_needed"], dropna=False)
        .size()
        .reset_index(name="arms")
    )

    xlsx_path = DATA_DIR / "RV_LV_data_v0.2_STRATIFY_locked.xlsx"
    csv_path = DATA_DIR / "RV_LV_data_v0.2_STRATIFY_locked.csv"
    md_path = DATA_DIR / "RV_LV_data_v0.2_STRATIFY_locked.md"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="RV_LV_Arm_Level_v0.2", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
    style_excel(xlsx_path)
    md_path.write_text(
        "\n".join(
            [
                "# RV/LV Data v0.2 STRATIFY Locked",
                "",
                "Incremental update from RV_LV_data_v0.1. STRATIFY 2026 added from full-text Table 2.",
                "",
                "## Summary",
                "",
                dataframe_to_markdown(summary),
                "",
                "## STRATIFY Rows",
                "",
                dataframe_to_markdown(df[df["study_id"] == "STRATIFY_2026"]),
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(f"WROTE {csv_path}")
    print(f"WROTE {xlsx_path}")
    print(f"WROTE {md_path}")


def write_audit_note() -> None:
    note = DATA_DIR / "STRATIFY_2026_fulltext_verification_v0.1.md"
    note.write_text(
        "\n".join(
            [
                "# STRATIFY 2026 Full-Text Verification v0.1",
                "",
                "Source PDF: `V1-ALL/16 stratify.pdf` in the project full-text PDF archive.",
                "",
                "Verified from full-text Results, Figure 1, and Table 2.",
                "",
                "Key locked values:",
                "",
                "- Population: acute intermediate-high-risk PE by ESC criteria.",
                "- Randomized: USAT 71, IV low-dose alteplase 70, heparin 69.",
                "- Death by 3 months: USAT 2, IV low-dose alteplase 4, heparin 0.",
                "- Full-dose thrombolysis for clinical deterioration: USAT 1, IV low-dose alteplase 2, heparin 2.",
                "- ISTH major bleeding: USAT 6, IV low-dose alteplase 6, heparin 1.",
                "- TIMI major bleeding: USAT 2, IV low-dose alteplase 2, heparin 0.",
                "- RV/LV diameter ratio reduction at 48-96h: USAT 0.3 SD 0.3, IV low-dose alteplase 0.4 SD 0.4, heparin 0.3 SD 0.3.",
                "",
                "Important caveat:",
                "",
                "- The article text reports two fatal intracranial haemorrhages overall, but Table 2 does not provide complete arm-specific intracranial haemorrhage counts. Therefore STRATIFY is marked not ready for ICH NMA unless supplementary data or author clarification becomes available.",
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(f"WROTE {note}")


def main() -> None:
    update_core_outcomes()
    update_rv_lv()
    write_audit_note()


if __name__ == "__main__":
    main()
