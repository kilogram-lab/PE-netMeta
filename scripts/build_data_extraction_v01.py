from __future__ import annotations

import pathlib

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTDIR = pathlib.Path("04_data_extraction")
OUTDIR.mkdir(parents=True, exist_ok=True)

DEFAULT_OUT = pathlib.Path(
    r"D:/001综合文件夹海康备份/44 自己的系列/52自己的写作文章/02肺栓塞和DVT/02肺栓塞网状meta-cdex版"
)


study_rows = [
    {
        "study_id": "TOPCOAT_2014",
        "short_name": "TOPCOAT",
        "year": 2014,
        "citation": "Kline et al. Treatment of submassive pulmonary embolism with tenecteplase or placebo",
        "dedup_status": "unique",
        "source_pdf": "03 Treatment of submassive pulmonary .pdf",
        "design": "multicenter double-blind placebo-controlled RCT",
        "population": "Normotensive submassive PE with right ventricular strain",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs biomarker extraction",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 83,
        "arm_n_summary": "Tenecteplase n=40; placebo/AC n=43",
        "follow_up": "3 months",
        "primary_outcome_definition": "Composite patient-oriented cardiopulmonary outcome at 3 months; includes death/circulatory shock/intubation/recurrent PE and functional/RV outcomes",
        "direct_nma_ready": "yes, for ST vs AC; outcome definitions need harmonization",
        "notes": "Original TOPCOAT main report; replaces earlier biomarker-only secondary report as primary source.",
    },
    {
        "study_id": "ULTIMA_2014",
        "short_name": "ULTIMA",
        "year": 2014,
        "citation": "Randomized controlled trial of ultrasound-assisted catheter-directed thrombolysis for acute intermediate-risk PE",
        "dedup_status": "unique",
        "source_pdf": "04 Randomized controlled trial of ultrasound assi.pdf",
        "design": "open-label randomized controlled trial",
        "population": "Acute intermediate-risk PE; RV/LV ratio >=1.0",
        "risk_mapping_v01": "intermediate-risk; likely NMA-3; NMA-1/2 needs troponin/biomarker extraction",
        "node_comparison": "USCDT vs AC",
        "total_n_randomized": 59,
        "arm_n_summary": "USCDT/USAT + UFH n=30; UFH alone n=29",
        "follow_up": "24 h primary imaging; clinical outcomes to 90 days",
        "primary_outcome_definition": "Change in RV/LV ratio from baseline to 24 h",
        "direct_nma_ready": "yes, for USCDT vs AC",
        "notes": "Key USCDT-AC direct edge.",
    },
    {
        "study_id": "PEITHO_2014",
        "short_name": "PEITHO",
        "year": 2014,
        "citation": "Meyer et al. Fibrinolysis for patients with intermediate-risk pulmonary embolism",
        "dedup_status": "duplicate source present; keep one study record",
        "source_pdf": "05 Meyer-2014-Fibrinolysis for patients with inte.pdf; previous 44 Fibrinolysis PDF",
        "design": "randomized double-blind placebo-controlled trial",
        "population": "Intermediate-risk PE with RV dysfunction and positive troponin; hemodynamically stable",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 1005,
        "arm_n_summary": "Tenecteplase + heparin n=506; placebo + heparin n=499",
        "follow_up": "7 days primary; 30 days safety/secondary",
        "primary_outcome_definition": "Death or hemodynamic decompensation/collapse within 7 days",
        "direct_nma_ready": "yes, for ST vs AC",
        "notes": "Core ST-AC large RCT.",
    },
    {
        "study_id": "SUNSET_sPE_2021",
        "short_name": "SUNSET sPE",
        "year": 2021,
        "citation": "Randomized trial comparing standard versus ultrasound-assisted thrombolysis for submassive PE",
        "dedup_status": "unique",
        "source_pdf": "09 Randomized Trial Comparing Standard Source JAC.pdf",
        "design": "multicenter randomized head-to-head single-blind trial",
        "population": "Submassive/intermediate-risk PE",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs biomarker extraction",
        "node_comparison": "CDT vs USCDT",
        "total_n_randomized": 81,
        "arm_n_summary": "USCDT/USAT vs standard CDT; exact arm n to verify from table",
        "follow_up": "48 h imaging primary; clinical follow-up to verify",
        "primary_outcome_definition": "Pulmonary arterial thrombus reduction",
        "direct_nma_ready": "yes, for CDT vs USCDT; clinical event sparsity expected",
        "notes": "Key CDT-USCDT direct edge.",
    },
    {
        "study_id": "CANARY_2022",
        "short_name": "CANARY",
        "year": 2022,
        "citation": "Catheter-directed thrombolysis vs anticoagulation in acute intermediate-high-risk PE",
        "dedup_status": "unique",
        "source_pdf": "11 Catheter Directed Thrombolysis vs An Source JA.pdf",
        "design": "open-label randomized clinical trial with masked endpoint adjudication",
        "population": "Acute intermediate-high-risk PE",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "CDT vs AC",
        "total_n_randomized": 94,
        "arm_n_summary": "cCDT + anticoagulation n=48; anticoagulation n=46",
        "follow_up": "72 h and 3 months",
        "primary_outcome_definition": "3-month RV/LV ratio >0.9",
        "direct_nma_ready": "yes, for CDT vs AC",
        "notes": "Prematurely terminated; 85 completed 3-month echo follow-up.",
    },
    {
        "study_id": "PEERLESS_2025",
        "short_name": "PEERLESS",
        "year": 2025,
        "citation": "Large-bore mechanical thrombectomy vs catheter-directed thrombolysis in intermediate-risk PE",
        "dedup_status": "unique",
        "source_pdf": "14 Large bore Mechanical Thrombectomy V Source Ci.pdf",
        "design": "prospective multicenter randomized controlled trial",
        "population": "Intermediate-risk PE with RV dilatation and additional clinical risk factors",
        "risk_mapping_v01": "intermediate-risk; likely NMA-3; NMA-1/2 needs biomarker/risk-factor mapping",
        "node_comparison": "LBAT vs CDT-mixed",
        "total_n_randomized": 550,
        "arm_n_summary": "FlowTriever LBMT n=274; CDT n=276",
        "follow_up": "through discharge or 7 days for primary endpoint; 30 days and later outcomes to verify",
        "primary_outcome_definition": "Hierarchical win-ratio composite: mortality, ICH, major bleeding, clinical deterioration/bailout, ICU use",
        "direct_nma_ready": "yes with caveat: CDT arm is mixed and includes USCDT and other CDT variants",
        "notes": "CDT arm was not standardized; node handling may require sensitivity analysis.",
    },
    {
        "study_id": "STORM_PE_2025",
        "short_name": "STORM-PE",
        "year": 2025,
        "citation": "Mechanical thrombectomy with anticoagulation vs anticoagulation alone for acute intermediate-high-risk PE",
        "dedup_status": "unique",
        "source_pdf": "15 Randomized Controlled Trial of Mecha Source Ci.pdf",
        "design": "international randomized controlled trial",
        "population": "Acute intermediate-high-risk PE; symptoms <=14 days; RV enlargement and elevated cardiac biomarkers",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "CAT vs AC",
        "total_n_randomized": 100,
        "arm_n_summary": "CAVT/Penumbra Lightning Flash 16F + AC n=47; AC alone n=53",
        "follow_up": "48 h primary imaging; 7 days safety; 90 days planned",
        "primary_outcome_definition": "Change in RV/LV ratio at 48 h",
        "direct_nma_ready": "yes, for CAT vs AC",
        "notes": "Computer-assisted vacuum thrombectomy mapped to CAT.",
    },
    {
        "study_id": "STRATIFY_2026",
        "short_name": "STRATIFY",
        "year": 2026,
        "citation": "Low-dose USAT or intravenous thrombolysis or heparin for intermediate-high-risk PE",
        "dedup_status": "unique",
        "source_pdf": "16 stratify.pdf",
        "design": "investigator-initiated open-label multicenter 3-arm randomized clinical trial",
        "population": "Acute intermediate-high-risk PE",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "USCDT vs ST-low-dose vs AC",
        "total_n_randomized": 210,
        "arm_n_summary": "USAT low-dose alteplase n=70; IV low-dose alteplase n=70; heparin n=70",
        "follow_up": "48-96 h imaging primary; clinical assessment at 3 months",
        "primary_outcome_definition": "Reduction in refined Modified Miller Score",
        "direct_nma_ready": "yes, with low-dose ST field retained",
        "notes": "Low-dose IV alteplase is mapped to ST in main node, with dose sensitivity flag.",
    },
    {
        "study_id": "PRETHA_2026",
        "short_name": "PRETHA",
        "year": 2026,
        "citation": "Percutaneous reperfusion therapies vs anticoagulation in acute intermediate-high-risk PE",
        "dedup_status": "unique",
        "source_pdf": "17 Percutaneous Reperfusion Therapies v Source Pulm Circ SO 2026.pdf",
        "design": "single-center prospective 3-arm randomized clinical trial",
        "population": "Acute intermediate-high-risk PE",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "CAT vs CDT vs AC",
        "total_n_randomized": 39,
        "arm_n_summary": "Penumbra Indigo CAT8 thrombectomy n=13; trans-catheter thrombolysis n to verify; anticoagulation n to verify",
        "follow_up": "48 h; 1, 6, and 12 months",
        "primary_outcome_definition": "Echocardiographic, hemodynamic, and biomarker changes",
        "direct_nma_ready": "yes, small 3-arm trial",
        "notes": "Thrombectomy uses Penumbra Indigo CAT8; mapped to CAT.",
    },
    {
        "study_id": "HI_PEITHO_2026",
        "short_name": "HI-PEITHO",
        "year": 2026,
        "citation": "Ultrasound-facilitated catheter-directed fibrinolysis plus anticoagulation vs anticoagulation alone",
        "dedup_status": "unique",
        "source_pdf": "18 Ultrasound Facilitated Catheter Dir Source N Engl J Med SO 2026.pdf",
        "design": "adaptive-design open-label randomized trial with blinded adjudication",
        "population": "Acute intermediate-risk PE with RV/LV and clinical-risk criteria",
        "risk_mapping_v01": "intermediate-risk; likely enriched intermediate-high; needs exact biomarker mapping",
        "node_comparison": "USCDT vs AC",
        "total_n_randomized": "NR in v0.1",
        "arm_n_summary": "USCDT + AC vs AC; exact n to extract from results table",
        "follow_up": "7 days primary; 30 days major bleeding/ICH",
        "primary_outcome_definition": "Composite pulmonary embolism-related death, cardiorespiratory decompensation/collapse, or nonfatal symptomatic recurrent PE within 7 days",
        "direct_nma_ready": "yes, key clinical-outcome USCDT-AC RCT",
        "notes": "Need extract exact denominators/events from results table.",
    },
    {
        "study_id": "MAPPET3_2002",
        "short_name": "MAPPET-3",
        "year": 2002,
        "citation": "Konstantinides et al. Heparin plus alteplase compared with heparin alone in submassive PE",
        "dedup_status": "unique",
        "source_pdf": "19 Heparin plus alteplase.PDF",
        "design": "randomized trial",
        "population": "Submassive PE",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs biomarker/RV criteria mapping",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 256,
        "arm_n_summary": "Heparin + alteplase n=118; heparin alone n=138",
        "follow_up": "in-hospital and 30 days to verify",
        "primary_outcome_definition": "Death or clinical deterioration requiring escalation of treatment",
        "direct_nma_ready": "yes, for ST vs AC",
        "notes": "Classic MAPPET-3 trial; exact event extraction needed from tables.",
    },
    {
        "study_id": "HAIRE_1993",
        "short_name": "Haire 1993",
        "year": 1993,
        "citation": "Alteplase versus heparin in acute pulmonary embolism",
        "dedup_status": "unique but text extraction insufficient",
        "source_pdf": "20 Alteplase versus heparin i.PDF",
        "design": "randomized trial reported, needs OCR/manual reading",
        "population": "acute PE; risk level unclear",
        "risk_mapping_v01": "pending",
        "node_comparison": "ST vs AC",
        "total_n_randomized": "NR",
        "arm_n_summary": "NR",
        "follow_up": "NR",
        "primary_outcome_definition": "NR",
        "direct_nma_ready": "pending OCR/manual verification",
        "notes": "Only 749 characters extracted; do not use for quantitative synthesis until OCR verified.",
    },
    {
        "study_id": "CDT_PILOT_2022",
        "short_name": "EuroIntervention pilot CDT",
        "year": 2022,
        "citation": "Pilot randomised trial of catheter-directed thrombolysis or standard anticoagulation for intermediate-high-risk PE",
        "dedup_status": "unique",
        "source_pdf": "21 A pilot randomised trial of catheter dire...PDF.pdf",
        "design": "pilot randomized study",
        "population": "Intermediate-high risk acute PE",
        "risk_mapping_v01": "intermediate-high risk; NMA-2 and NMA-3",
        "node_comparison": "CDT vs AC",
        "total_n_randomized": 23,
        "arm_n_summary": "CDT n=12; standard anticoagulation n=11",
        "follow_up": "24 h/48 h imaging; later follow-up to verify",
        "primary_outcome_definition": "Improvement of RV function, clot burden, and normotension at 24 h",
        "direct_nma_ready": "yes, small pilot CDT-AC trial",
        "notes": "Envelope randomization; small sample.",
    },
    {
        "study_id": "FASULLO_2011",
        "short_name": "Fasullo 2011",
        "year": 2011,
        "citation": "Six-month echocardiographic study in submassive PE and RV dysfunction",
        "dedup_status": "duplicate source present; keep one study record",
        "source_pdf": "27 Six month echocardiographic study...PDF",
        "design": "randomized trial",
        "population": "Submassive PE with RV dysfunction and normal blood pressure",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs biomarker extraction",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 72,
        "arm_n_summary": "Thrombolysis n=37; heparin/placebo n=35",
        "follow_up": "6 months",
        "primary_outcome_definition": "RV function/PASP and echocardiographic recovery",
        "direct_nma_ready": "yes, for ST vs AC",
        "notes": "Duplicate with prior screening batch.",
    },
    {
        "study_id": "SINHA_2017",
        "short_name": "Sinha 2017",
        "year": 2017,
        "citation": "Efficacy and safety of thrombolytic therapy in acute submassive PE",
        "dedup_status": "duplicate source present; keep one study record",
        "source_pdf": "28 Efficacy and Safety of Thrombolytic...pdf",
        "design": "prospective randomized single-center study",
        "population": "Acute submassive/intermediate-risk PE",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs exact criteria mapping",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 87,
        "arm_n_summary": "Tenecteplase + UFH vs placebo + UFH; exact arm n to verify",
        "follow_up": "in-hospital/short-term; exact follow-up to verify",
        "primary_outcome_definition": "Death, hemodynamic decompensation, recurrent PE, major bleeding, hemorrhagic stroke",
        "direct_nma_ready": "yes, for ST vs AC",
        "notes": "Duplicate with prior screening batch.",
    },
    {
        "study_id": "ZHANG_2018",
        "short_name": "Zhang 2018",
        "year": 2018,
        "citation": "Low-dose recombinant tissue-type plasminogen activator for acute intermediate-risk PE",
        "dedup_status": "duplicate source present; keep one study record",
        "source_pdf": "31 Clinical efficacy of low dose recombinant tissue-type plasminogen activator...pdf",
        "design": "randomized trial",
        "population": "Acute intermediate-risk PE",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs exact criteria mapping",
        "node_comparison": "ST-low-dose vs AC",
        "total_n_randomized": 66,
        "arm_n_summary": "Low-dose rt-PA n=33; LMWH n=33",
        "follow_up": "short-term and clinical follow-up to verify",
        "primary_outcome_definition": "Clinical efficacy/safety, RV/PAP measures",
        "direct_nma_ready": "yes, for ST vs AC with low-dose flag",
        "notes": "Duplicate with prior screening batch.",
    },
    {
        "study_id": "MOPETT_2013",
        "short_name": "MOPETT",
        "year": 2013,
        "citation": "Moderate pulmonary embolism treated with thrombolysis",
        "dedup_status": "unique",
        "source_pdf": "02 Moderate Pulmonary Embolism.pdf",
        "design": "prospective controlled randomized single-center open study",
        "population": "Moderate PE",
        "risk_mapping_v01": "needs mapping; may be intermediate-risk or sensitivity-only",
        "node_comparison": "ST-low-dose vs AC",
        "total_n_randomized": 121,
        "arm_n_summary": "safe-dose tPA + AC n=61; AC n=60",
        "follow_up": "28 months",
        "primary_outcome_definition": "Pulmonary hypertension and composite of pulmonary hypertension plus recurrent PE",
        "direct_nma_ready": "pending risk mapping; likely sensitivity if moderate PE cannot map cleanly",
        "notes": "No bleeding reported; composite death plus recurrent PE 1.6% vs 10%.",
    },
    {
        "study_id": "AHMED_2018",
        "short_name": "Ahmed 2018",
        "year": 2018,
        "citation": "Value of thrombolytic therapy for submassive pulmonary embolism",
        "dedup_status": "from prior screening batch; not in V1-ALL",
        "source_pdf": "49 value_of_thrombolytic_therapy_for_submassive.13(1).pdf",
        "design": "randomized controlled trial",
        "population": "Submassive PE, hemodynamically stable; RV/biomarker details to verify",
        "risk_mapping_v01": "intermediate-risk; NMA-3; NMA-1/2 needs extraction",
        "node_comparison": "ST vs AC",
        "total_n_randomized": 52,
        "arm_n_summary": "Streptokinase + anticoagulation vs anticoagulation; exact arm n to verify",
        "follow_up": "short-term; exact follow-up to verify",
        "primary_outcome_definition": "Clinical and echo/hemodynamic outcomes; exact definition to extract",
        "direct_nma_ready": "yes after exact outcome extraction",
        "notes": "Included from previous formal screening v0.1.",
    },
]


arm_rows = [
    # study_id, arm_label, node, n, death, clinical, major_bleed, ich, recurrent_pe, rvlv, follow-up, notes
    ["TOPCOAT_2014", "Tenecteplase", "ST", 40, "1 fatal ICH-related death? verify", "Composite adverse outcome: fewer than placebo; exact event count to verify", "hemorrhagic events reported; exact major count to verify", "1 fatal ICH", "included in composite; exact count to verify", "RV strain/function at 3 months; exact values to extract", "3 months", "Do not use mortality count until table checked"],
    ["TOPCOAT_2014", "Placebo + anticoagulation", "AC", 43, "0? verify", "Composite adverse outcomes: 3 placebo-treated patients in abstract", "NR", "0 reported in abstract", "included in composite; exact count to verify", "RV strain/function at 3 months; exact values to extract", "3 months", "Original TOPCOAT"],
    ["ULTIMA_2014", "USAT/USCDT + UFH", "USCDT", 30, "0", "NR", "0", "0", "0 recurrent VTE", "RV/LV 1.28±0.19 to 0.99±0.17 at 24h", "90 days", "Primary RV/LV change"],
    ["ULTIMA_2014", "UFH alone", "AC", 29, "1", "NR", "0", "0", "0 recurrent VTE", "RV/LV 1.20±0.14 to 1.17±0.20 at 24h", "90 days", "One death in heparin group"],
    ["PEITHO_2014", "Tenecteplase + heparin", "ST", 506, "6 by day 7; verify table", "Hemodynamic decompensation/collapse 8; primary composite 13", "Major extracranial bleeding 32", "ICH 10; stroke 12", "NR", "NR", "7/30 days", "Counts from known PEITHO report; recheck against PDF table before analysis"],
    ["PEITHO_2014", "Placebo + heparin", "AC", 499, "9 by day 7; verify table", "Hemodynamic decompensation/collapse 25; primary composite 28", "Major extracranial bleeding 6", "ICH 1; stroke 1", "NR", "NR", "7/30 days", "Counts from known PEITHO report; recheck against PDF table before analysis"],
    ["SUNSET_sPE_2021", "USAT/EKOS", "USCDT", "approx 40; verify", "NR", "NR", "NR", "NR", "NR", "Pulmonary arterial thrombus reduction; exact RV/LV to extract", "48 h primary", "Head-to-head imaging endpoint"],
    ["SUNSET_sPE_2021", "Standard CDT", "CDT", "approx 41; verify", "NR", "NR", "NR", "NR", "NR", "Pulmonary arterial thrombus reduction; exact RV/LV to extract", "48 h primary", "Head-to-head imaging endpoint"],
    ["CANARY_2022", "cCDT + anticoagulation", "CDT", 48, "3-month mortality included as secondary; exact count to verify", "Death or RV/LV >0.9 composite lower; exact full line to verify", "1 major GI bleeding", "NR", "NR", "RV/LV >0.9 at 72h: 13/48; at 3mo primary: 2/46", "72 h/3 months", "Completed 3-month echo n=85"],
    ["CANARY_2022", "Anticoagulation", "AC", 46, "3-month mortality included as secondary; exact count to verify", "Death or RV/LV >0.9 composite; exact full line to verify", "0? verify", "NR", "NR", "RV/LV >0.9 at 72h: 24/46; at 3mo primary: 5/39", "72 h/3 months", "Completed 3-month echo n=85"],
    ["PEERLESS_2025", "FlowTriever LBMT", "LBAT", 274, "No significant difference vs CDT; exact count to extract", "Clinical deterioration/bailout 1.8%", "No significant difference; exact count to extract", "No significant difference; exact count to extract", "NR", "RV/LV ratio reduction 0.32±0.24", "discharge/7 days", "Primary win-ratio composite"],
    ["PEERLESS_2025", "CDT mixed", "CDT", 276, "No significant difference vs LBMT; exact count to extract", "Clinical deterioration/bailout 5.4%", "No significant difference; exact count to extract", "No significant difference; exact count to extract", "NR", "RV/LV ratio reduction 0.30±0.26", "discharge/7 days", "CDT arm includes USCDT/standard/pharmacomechanical CDT; sensitivity needed"],
    ["STORM_PE_2025", "CAVT/Penumbra + AC", "CAT", 47, "NR", "NR", "Safety outcomes at 7 days; exact count to extract", "NR", "NR", "Primary: RV/LV ratio change at 48h; exact value to extract", "48 h/7 d/90 d", "Penumbra Lightning Flash 16F"],
    ["STORM_PE_2025", "AC alone", "AC", 53, "NR", "NR", "Safety outcomes at 7 days; exact count to extract", "NR", "NR", "Primary: RV/LV ratio change at 48h; exact value to extract", "48 h/7 d/90 d", "Intermediate-high risk"],
    ["STRATIFY_2026", "USAT low-dose alteplase", "USCDT", 70, "Death increased with low-dose thrombolysis; exact count to extract", "Clinical endpoints assessed; exact count to extract", "Bleeding numerically more frequent; exact count to extract", "NR", "NR", "rmMS reduction; RV/LV not primary", "48-96 h/3 mo", "20 mg alteplase via USAT"],
    ["STRATIFY_2026", "IV low-dose alteplase", "ST", 70, "Death increased with low-dose thrombolysis; exact count to extract", "Clinical endpoints assessed; exact count to extract", "Bleeding numerically more frequent; exact count to extract", "NR", "NR", "rmMS reduction; RV/LV not primary", "48-96 h/3 mo", "20 mg IV alteplase; low-dose flag"],
    ["STRATIFY_2026", "Heparin alone", "AC", 70, "Exact count to extract", "Clinical endpoints assessed; exact count to extract", "Exact count to extract", "NR", "NR", "rmMS reduction comparator", "48-96 h/3 mo", "AC comparator"],
    ["PRETHA_2026", "Penumbra Indigo CAT8 thrombectomy", "CAT", 13, "NR", "NR", "Adverse events 1/13? verify event type", "NR", "NR", "RV/LV decreased by 0.3; PAP decreased 29%", "48 h/12 mo", "Small 3-arm trial"],
    ["PRETHA_2026", "Trans-catheter thrombolysis", "CDT", "to verify", "NR", "NR", "Adverse events highest 38%; exact denominator to verify", "NR", "NR", "RV/LV decreased by 0.4; PAP decreased 39%", "48 h/12 mo", "Need arm n from table"],
    ["PRETHA_2026", "Anticoagulation", "AC", "to verify", "NR", "NR", "NR", "NR", "NR", "Comparator for echo/hemodynamic changes", "48 h/12 mo", "Need arm n from table"],
    ["HI_PEITHO_2026", "USCDT + AC", "USCDT", "to extract", "Primary composite 11 patients? confirm arm and denominator", "Primary composite within 7d; exact count to extract", "30d major bleeding exact count to extract", "No ICH reported", "Primary includes nonfatal symptomatic recurrent PE", "NR", "7/30 days", "Large clinical-outcome RCT"],
    ["HI_PEITHO_2026", "AC alone", "AC", "to extract", "Primary composite exact count to extract", "Primary composite within 7d; exact count to extract", "30d major bleeding exact count to extract", "No ICH reported", "Primary includes nonfatal symptomatic recurrent PE", "NR", "7/30 days", "Large clinical-outcome RCT"],
    ["MAPPET3_2002", "Heparin + alteplase", "ST", 118, "Exact count to extract", "Death or clinical deterioration; exact count to extract", "Exact count to extract", "Exact count to extract", "Exact count to extract", "NR", "in-hospital/30d", "Classic submassive PE trial"],
    ["MAPPET3_2002", "Heparin alone", "AC", 138, "Exact count to extract", "Death or clinical deterioration; exact count to extract", "Exact count to extract", "Exact count to extract", "Exact count to extract", "NR", "in-hospital/30d", "Classic submassive PE trial"],
    ["HAIRE_1993", "Alteplase", "ST", "NR", "NR", "NR", "NR", "NR", "NR", "NR", "OCR needed", "Do not analyze yet"],
    ["HAIRE_1993", "Heparin", "AC", "NR", "NR", "NR", "NR", "NR", "NR", "NR", "OCR needed", "Do not analyze yet"],
    ["CDT_PILOT_2022", "CDT alteplase 20 mg", "CDT", 12, "NR", "Primary efficacy endpoint achieved more frequently; exact count to extract", "NR", "NR", "NR", "RV function/clot burden/normotension endpoint", "24/48 h", "Pilot RCT"],
    ["CDT_PILOT_2022", "Standard anticoagulation", "AC", 11, "NR", "Primary efficacy endpoint lower; exact count to extract", "NR", "NR", "NR", "RV function/clot burden/normotension endpoint", "24/48 h", "Pilot RCT"],
    ["FASULLO_2011", "Thrombolysis + heparin", "ST", 37, "Exact count to extract", "Clinical worsening/echo outcomes to extract", "Exact count to extract", "NR", "Exact count to extract", "RV/PASP outcomes at 6 months", "6 months", "Duplicate source from prior batch"],
    ["FASULLO_2011", "Heparin/placebo", "AC", 35, "Exact count to extract", "Clinical worsening/echo outcomes to extract", "Exact count to extract", "NR", "Exact count to extract", "RV/PASP outcomes at 6 months", "6 months", "Duplicate source from prior batch"],
    ["SINHA_2017", "Tenecteplase + UFH", "ST", "to verify", "Exact count to extract", "Hemodynamic decompensation; exact count to extract", "Exact count to extract", "Hemorrhagic stroke; exact count to extract", "Exact count to extract", "NR", "short-term", "Arm n to verify"],
    ["SINHA_2017", "Placebo + UFH", "AC", "to verify", "Exact count to extract", "Hemodynamic decompensation; exact count to extract", "Exact count to extract", "Hemorrhagic stroke; exact count to extract", "Exact count to extract", "NR", "short-term", "Arm n to verify"],
    ["ZHANG_2018", "Low-dose rt-PA + LMWH", "ST", 33, "Exact count to extract", "Clinical efficacy outcome; exact count to extract", "Exact count to extract", "Exact count to extract", "Exact count to extract", "RV/PAP outcomes to extract", "to verify", "Low-dose ST flag"],
    ["ZHANG_2018", "LMWH", "AC", 33, "Exact count to extract", "Clinical efficacy outcome; exact count to extract", "Exact count to extract", "Exact count to extract", "Exact count to extract", "RV/PAP outcomes to extract", "to verify", "Low-dose ST comparator"],
    ["MOPETT_2013", "Safe-dose tPA + AC", "ST", 61, "No significant difference; exact count to extract", "Pulmonary hypertension 9/61; PH+recurrent PE composite exact table", "0", "0", "No recurrent PE reported in TG? verify table", "PASP/PH endpoint", "28 months", "Risk mapping pending"],
    ["MOPETT_2013", "AC alone", "AC", 60, "No significant difference; exact count to extract", "Pulmonary hypertension 32/60; PH+recurrent PE composite exact table", "0", "0", "recurrent PE exact count to extract", "PASP/PH endpoint", "28 months", "Risk mapping pending"],
    ["AHMED_2018", "Streptokinase + anticoagulation", "ST", "to verify", "Exact count to extract", "Exact count to extract", "Exact count to extract", "NR", "Exact count to extract", "RVD/PAP/biomarker to extract", "to verify", "Prior batch"],
    ["AHMED_2018", "Anticoagulation", "AC", "to verify", "Exact count to extract", "Exact count to extract", "Exact count to extract", "NR", "Exact count to extract", "RVD/PAP/biomarker to extract", "to verify", "Prior batch"],
]


arm_columns = [
    "study_id",
    "arm_label",
    "node",
    "n_randomized",
    "death",
    "clinical_deterioration_or_primary_clinical",
    "major_bleeding",
    "intracranial_hemorrhage",
    "recurrent_pe",
    "rv_lv_or_rv_recovery",
    "follow_up",
    "notes",
]


dict_rows = [
    {"field": "NR", "meaning": "Not reported or not reliably extracted in v0.1; requires full table extraction/OCR"},
    {"field": "to verify", "meaning": "Likely available in PDF but not locked in this v0.1 extraction pass"},
    {"field": "ST", "meaning": "Systemic thrombolysis, including low-dose systemic alteplase unless flagged"},
    {"field": "CDT", "meaning": "Catheter-directed thrombolysis without ultrasound as main mechanism"},
    {"field": "USCDT", "meaning": "Ultrasound-assisted/facilitated catheter-directed thrombolysis"},
    {"field": "LBAT", "meaning": "Large-bore aspiration/mechanical thrombectomy, e.g. FlowTriever"},
    {"field": "CAT", "meaning": "Catheter-assisted aspiration thrombectomy, e.g. Penumbra Indigo/Lightning"},
    {"field": "AC", "meaning": "Anticoagulation/control medical therapy"},
]


def style_excel(path: pathlib.Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = {
            "A": 18,
            "B": 22,
            "C": 12,
            "D": 48,
            "E": 18,
            "F": 36,
            "G": 30,
            "H": 30,
            "I": 18,
            "J": 24,
            "K": 18,
            "L": 60,
            "M": 32,
            "N": 22,
        }
        for idx in range(1, ws.max_column + 1):
            col = get_column_letter(idx)
            ws.column_dimensions[col].width = widths.get(col, 24)
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)


def dataframe_to_markdown(df: pd.DataFrame) -> str:
    columns = [str(column) for column in df.columns]
    rows = [[str(value).replace("\n", " ") for value in row] for row in df.itertuples(index=False, name=None)]
    lines = [
        "| " + " | ".join(columns) + " |",
        "| " + " | ".join(["---"] * len(columns)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


def main() -> None:
    studies = pd.DataFrame(study_rows)
    arms = pd.DataFrame(arm_rows, columns=arm_columns)
    dictionary = pd.DataFrame(dict_rows)

    xlsx_path = OUTDIR / "data_extraction_v0.1.xlsx"
    csv_studies = OUTDIR / "data_extraction_studies_v0.1.csv"
    csv_arms = OUTDIR / "data_extraction_arms_v0.1.csv"
    md_path = OUTDIR / "data_extraction_v0.1.md"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        studies.to_excel(writer, sheet_name="Study_Level", index=False)
        arms.to_excel(writer, sheet_name="Arm_Level", index=False)
        dictionary.to_excel(writer, sheet_name="Data_Dictionary", index=False)

    style_excel(xlsx_path)
    studies.to_csv(csv_studies, index=False, encoding="utf-8-sig")
    arms.to_csv(csv_arms, index=False, encoding="utf-8-sig")

    md = [
        "# Data Extraction v0.1",
        "",
        "This is a first formal extraction table for the PE network meta-analysis. Values marked `NR`, `to verify`, or `待核验` must not be used as locked quantitative inputs until checked against the original article tables.",
        "",
        "## Study-Level Table",
        "",
        dataframe_to_markdown(studies),
        "",
        "## Arm-Level Table",
        "",
        dataframe_to_markdown(arms),
        "",
        "## Data Dictionary",
        "",
        dataframe_to_markdown(dictionary),
        "",
    ]
    md_path.write_text("\n".join(md), encoding="utf-8")

    DEFAULT_OUT.mkdir(parents=True, exist_ok=True)
    for src, name in [
        (xlsx_path, "正式数据提取表_v0.1.xlsx"),
        (csv_studies, "正式数据提取表_研究层_v0.1.csv"),
        (csv_arms, "正式数据提取表_治疗臂层_v0.1.csv"),
        (md_path, "正式数据提取表_v0.1.md"),
    ]:
        (DEFAULT_OUT / name).write_bytes(src.read_bytes())

    print(f"WROTE\t{xlsx_path}")
    print(f"WROTE\t{csv_studies}")
    print(f"WROTE\t{csv_arms}")
    print(f"WROTE\t{md_path}")
    print(f"COPIED_TO\t{DEFAULT_OUT}")
    print(f"STUDIES\t{len(studies)}")
    print(f"ARMS\t{len(arms)}")


if __name__ == "__main__":
    main()
